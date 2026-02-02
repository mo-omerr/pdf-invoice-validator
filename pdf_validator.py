#!/usr/bin/env python3
"""
AI-Powered PDF Invoice Validator with Multi-Template Support
- Automatically detects vendor from PDF
- Creates new templates for unknown vendors
- Validates and extracts invoice data
- Exports to Excel (one workbook per PDF, one sheet per invoice)
"""

import json
import base64
import sys
import os
import re
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Tuple

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv is optional

try:
    import anthropic
except ImportError:
    print("Error: anthropic is required. Install with: pip install anthropic")
    sys.exit(1)

try:
    import fitz  # PyMuPDF
except ImportError:
    print("Error: PyMuPDF is required. Install with: pip install PyMuPDF")
    sys.exit(1)

try:
    from PIL import Image
except ImportError:
    print("Error: Pillow is required. Install with: pip install Pillow")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Data Classes
# ============================================================================

@dataclass
class ValidationResult:
    """Stores validation results for a single invoice."""
    invoice_number: Optional[str] = None
    page_numbers: list = field(default_factory=list)
    is_valid: bool = True
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    extracted_data: dict = field(default_factory=dict)

    def add_error(self, error: str):
        self.errors.append(error)
        self.is_valid = False

    def add_warning(self, warning: str):
        self.warnings.append(warning)


@dataclass
class PDFValidationReport:
    """Stores overall validation report for the PDF."""
    filename: str
    template_name: str = ""
    template_created: bool = False
    total_pages: int = 0
    invoices_found: int = 0
    invoices_valid: int = 0
    invoices_invalid: int = 0
    invoice_results: list = field(default_factory=list)
    global_errors: list = field(default_factory=list)
    is_valid: bool = True

    def add_invoice_result(self, result: ValidationResult):
        self.invoice_results.append(result)
        if result.is_valid:
            self.invoices_valid += 1
        else:
            self.invoices_invalid += 1
            self.is_valid = False
        self.invoices_found += 1


# ============================================================================
# Template Manager
# ============================================================================

class TemplateManager:
    """Manages invoice templates for different vendors."""

    def __init__(self, templates_dir: str = None):
        if templates_dir is None:
            templates_dir = Path(__file__).parent / "templates"
        self.templates_dir = Path(templates_dir)
        self.templates_dir.mkdir(exist_ok=True)
        self.templates = {}
        self._load_all_templates()

    def _load_all_templates(self):
        """Load all templates from the templates directory."""
        for template_file in self.templates_dir.glob("*.json"):
            try:
                with open(template_file, 'r') as f:
                    template = json.load(f)
                    vendor_name = template.get("vendor", {}).get("name", "")
                    if vendor_name:
                        self.templates[vendor_name.lower()] = {
                            "path": str(template_file),
                            "data": template
                        }
                        print(f"Loaded template: {vendor_name}")
            except Exception as e:
                print(f"Warning: Could not load template {template_file}: {e}")

    def get_template_by_vendor(self, vendor_name: str) -> Optional[dict]:
        """Get template by vendor name (case-insensitive partial match)."""
        vendor_lower = vendor_name.lower()

        # Exact match first
        if vendor_lower in self.templates:
            return self.templates[vendor_lower]["data"]

        # Partial match
        for key, value in self.templates.items():
            if key in vendor_lower or vendor_lower in key:
                return value["data"]

        return None

    def get_all_vendor_names(self) -> list:
        """Get list of all known vendor names."""
        return [t["data"]["vendor"]["name"] for t in self.templates.values()]

    def save_template(self, template: dict) -> str:
        """Save a new template to the templates directory."""
        vendor_name = template.get("vendor", {}).get("name", "unknown")
        # Create safe filename
        safe_name = re.sub(r'[^\w\s-]', '', vendor_name.lower())
        safe_name = re.sub(r'[-\s]+', '_', safe_name)

        template_path = self.templates_dir / f"{safe_name}.json"

        with open(template_path, 'w') as f:
            json.dump(template, f, indent=2)

        # Add to loaded templates
        self.templates[vendor_name.lower()] = {
            "path": str(template_path),
            "data": template
        }

        print(f"Template saved: {template_path}")
        return str(template_path)


# ============================================================================
# Excel Exporter
# ============================================================================

class ExcelExporter:
    """Exports invoice data to Excel workbooks."""

    def __init__(self):
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.currency_format = '$#,##0.00'
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, color="FFFFFF")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_report_to_excel(self, report: PDFValidationReport, output_path: str = None) -> str:
        """Export validation report to Excel."""
        if output_path is None:
            pdf_name = Path(report.filename).stem
            output_path = f"{pdf_name}_invoices.xlsx"

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Create a sheet for each invoice
        for i, inv_result in enumerate(report.invoice_results, 1):
            inv_num = inv_result.invoice_number or f"Invoice_{i}"
            sheet_name = f"Invoice_{inv_num}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            self._populate_invoice_sheet(ws, inv_result)

        # Create summary sheet at the beginning
        summary_ws = wb.create_sheet(title="Summary", index=0)
        self._populate_summary_sheet(summary_ws, report)

        # Create dashboard sheet with charts
        dashboard_ws = wb.create_sheet(title="Dashboard", index=0)
        self._populate_dashboard_sheet(dashboard_ws, report)

        wb.save(output_path)
        return output_path

    def _populate_summary_sheet(self, ws, report: PDFValidationReport):
        """Populate the summary sheet with data and charts."""
        ws['A1'] = "Invoice Summary Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A3'] = "Source PDF:"
        ws['B3'] = report.filename
        ws['A4'] = "Template Used:"
        ws['B4'] = report.template_name
        ws['A5'] = "Template Created:"
        ws['B5'] = "Yes" if report.template_created else "No (Existing)"
        ws['A6'] = "Total Pages:"
        ws['B6'] = report.total_pages
        ws['A7'] = "Invoices Found:"
        ws['B7'] = report.invoices_found
        ws['A8'] = "Valid Invoices:"
        ws['B8'] = report.invoices_valid
        ws['A9'] = "Invalid Invoices:"
        ws['B9'] = report.invoices_invalid

        for row in range(3, 10):
            ws[f'A{row}'].font = Font(bold=True)

        # Invoice list header
        headers = ["Invoice #", "Date of Issue", "Due Date", "Billed To", "Amount Due", "Subtotal", "Tax", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')

        # Invoice list data
        data_start_row = 12
        total_amount = 0
        total_subtotal = 0
        total_tax = 0

        for row_idx, inv_result in enumerate(report.invoice_results, data_start_row):
            data = inv_result.extracted_data
            ws.cell(row=row_idx, column=1, value=inv_result.invoice_number or "Unknown").border = self.thin_border
            ws.cell(row=row_idx, column=2, value=data.get("date_of_issue", "")).border = self.thin_border
            ws.cell(row=row_idx, column=3, value=data.get("due_date", "")).border = self.thin_border
            ws.cell(row=row_idx, column=4, value=data.get("billed_to", "")).border = self.thin_border

            # Parse amount values (remove $ sign if present)
            amount_due = self._parse_currency(data.get("amount_due", 0))
            subtotal = self._parse_currency(data.get("subtotal", 0))
            tax = self._parse_currency(data.get("tax", 0))

            total_amount += amount_due
            total_subtotal += subtotal
            total_tax += tax

            amount_cell = ws.cell(row=row_idx, column=5, value=amount_due)
            amount_cell.number_format = self.currency_format
            amount_cell.border = self.thin_border

            subtotal_cell = ws.cell(row=row_idx, column=6, value=subtotal)
            subtotal_cell.number_format = self.currency_format
            subtotal_cell.border = self.thin_border

            tax_cell = ws.cell(row=row_idx, column=7, value=tax)
            tax_cell.number_format = self.currency_format
            tax_cell.border = self.thin_border

            status_cell = ws.cell(row=row_idx, column=8, value="VALID" if inv_result.is_valid else "INVALID")
            status_cell.border = self.thin_border
            if not inv_result.is_valid:
                status_cell.font = Font(color="FF0000")

        data_end_row = data_start_row + len(report.invoice_results) - 1

        # Totals row
        totals_row = data_end_row + 1
        ws.cell(row=totals_row, column=4, value="TOTALS:").font = Font(bold=True)

        total_amount_cell = ws.cell(row=totals_row, column=5, value=total_amount)
        total_amount_cell.number_format = self.currency_format
        total_amount_cell.font = Font(bold=True)
        total_amount_cell.border = self.thin_border

        total_subtotal_cell = ws.cell(row=totals_row, column=6, value=total_subtotal)
        total_subtotal_cell.number_format = self.currency_format
        total_subtotal_cell.font = Font(bold=True)
        total_subtotal_cell.border = self.thin_border

        total_tax_cell = ws.cell(row=totals_row, column=7, value=total_tax)
        total_tax_cell.number_format = self.currency_format
        total_tax_cell.font = Font(bold=True)
        total_tax_cell.border = self.thin_border

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # Add charts if we have data
        if len(report.invoice_results) > 0:
            chart_start_row = totals_row + 3

            # Chart 1: Amount Due Trend (Bar Chart)
            ws.cell(row=chart_start_row, column=1, value="Invoice Amount Trend").font = Font(bold=True, size=14)

            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.style = 10
            bar_chart.title = "Invoice Amounts by Date"
            bar_chart.y_axis.title = "Amount ($)"
            bar_chart.x_axis.title = "Invoice"

            data_ref = Reference(ws, min_col=5, min_row=11, max_row=data_end_row, max_col=5)
            cats_ref = Reference(ws, min_col=2, min_row=12, max_row=data_end_row)
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            bar_chart.shape = 4
            bar_chart.width = 15
            bar_chart.height = 10

            ws.add_chart(bar_chart, f"A{chart_start_row + 1}")

            # Chart 2: Subtotal vs Tax (Stacked Bar)
            ws.cell(row=chart_start_row, column=10, value="Subtotal vs Tax Breakdown").font = Font(bold=True, size=14)

            stacked_chart = BarChart()
            stacked_chart.type = "col"
            stacked_chart.grouping = "stacked"
            stacked_chart.style = 10
            stacked_chart.title = "Subtotal vs Tax per Invoice"
            stacked_chart.y_axis.title = "Amount ($)"

            data_ref2 = Reference(ws, min_col=6, min_row=11, max_row=data_end_row, max_col=7)
            cats_ref2 = Reference(ws, min_col=1, min_row=12, max_row=data_end_row)
            stacked_chart.add_data(data_ref2, titles_from_data=True)
            stacked_chart.set_categories(cats_ref2)
            stacked_chart.width = 15
            stacked_chart.height = 10

            ws.add_chart(stacked_chart, f"J{chart_start_row + 1}")

            # Chart 3: Line Chart showing trend over time
            line_chart_row = chart_start_row + 22
            ws.cell(row=line_chart_row, column=1, value="Amount Trend Over Time").font = Font(bold=True, size=14)

            line_chart = LineChart()
            line_chart.style = 10
            line_chart.title = "Invoice Amount Trend"
            line_chart.y_axis.title = "Amount ($)"
            line_chart.x_axis.title = "Invoice Date"

            data_ref3 = Reference(ws, min_col=5, min_row=11, max_row=data_end_row)
            cats_ref3 = Reference(ws, min_col=2, min_row=12, max_row=data_end_row)
            line_chart.add_data(data_ref3, titles_from_data=True)
            line_chart.set_categories(cats_ref3)
            line_chart.width = 15
            line_chart.height = 10

            ws.add_chart(line_chart, f"A{line_chart_row + 1}")

            # Chart 4: Pie chart for cost breakdown (if multiple invoices)
            if len(report.invoice_results) > 1:
                ws.cell(row=line_chart_row, column=10, value="Cost Distribution").font = Font(bold=True, size=14)

                pie_chart = PieChart()
                pie_chart.title = "Invoice Distribution"

                data_ref4 = Reference(ws, min_col=5, min_row=12, max_row=data_end_row)
                cats_ref4 = Reference(ws, min_col=1, min_row=12, max_row=data_end_row)
                pie_chart.add_data(data_ref4)
                pie_chart.set_categories(cats_ref4)
                pie_chart.width = 12
                pie_chart.height = 10

                # Add data labels
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.dataLabels.showVal = False
                pie_chart.dataLabels.showCatName = True

                ws.add_chart(pie_chart, f"J{line_chart_row + 1}")

    def _parse_currency(self, value):
        """Parse currency string to float."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Remove $ sign and commas
            cleaned = value.replace('$', '').replace(',', '').strip()
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0

    def _populate_dashboard_sheet(self, ws, report: PDFValidationReport):
        """Create a dashboard with key metrics and charts."""
        # Title
        ws['A1'] = "INVOICE DASHBOARD"
        ws['A1'].font = Font(bold=True, size=20)
        ws.merge_cells('A1:F1')

        # Key Metrics Section
        ws['A3'] = "KEY METRICS"
        ws['A3'].font = Font(bold=True, size=14)

        # Calculate metrics
        total_amount = 0
        total_subtotal = 0
        total_tax = 0
        amounts = []
        dates = []

        for inv_result in report.invoice_results:
            data = inv_result.extracted_data
            amount = self._parse_currency(data.get("amount_due", 0))
            subtotal = self._parse_currency(data.get("subtotal", 0))
            tax = self._parse_currency(data.get("tax", 0))

            total_amount += amount
            total_subtotal += subtotal
            total_tax += tax
            amounts.append(amount)
            dates.append(data.get("date_of_issue", ""))

        avg_amount = total_amount / len(report.invoice_results) if report.invoice_results else 0
        max_amount = max(amounts) if amounts else 0
        min_amount = min(amounts) if amounts else 0

        # Metrics cards
        metrics = [
            ("Total Invoices", report.invoices_found),
            ("Total Amount", f"${total_amount:,.2f}"),
            ("Total Subtotal", f"${total_subtotal:,.2f}"),
            ("Total Tax", f"${total_tax:,.2f}"),
            ("Average Invoice", f"${avg_amount:,.2f}"),
            ("Highest Invoice", f"${max_amount:,.2f}"),
            ("Lowest Invoice", f"${min_amount:,.2f}"),
            ("Vendor", report.template_name),
        ]

        # Create metric cards in a row
        col = 1
        for label, value in metrics[:4]:
            cell = ws.cell(row=5, column=col, value=label)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

            value_cell = ws.cell(row=6, column=col, value=value)
            value_cell.font = Font(bold=True, size=14)
            value_cell.alignment = Alignment(horizontal='center')
            col += 1

        col = 1
        for label, value in metrics[4:]:
            cell = ws.cell(row=8, column=col, value=label)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

            value_cell = ws.cell(row=9, column=col, value=value)
            value_cell.font = Font(bold=True, size=14)
            value_cell.alignment = Alignment(horizontal='center')
            col += 1

        # Data table for charts (hidden area)
        data_start_row = 12
        ws.cell(row=data_start_row, column=1, value="Date")
        ws.cell(row=data_start_row, column=2, value="Amount")
        ws.cell(row=data_start_row, column=3, value="Subtotal")
        ws.cell(row=data_start_row, column=4, value="Tax")
        ws.cell(row=data_start_row, column=5, value="Invoice #")

        for i, inv_result in enumerate(report.invoice_results):
            data = inv_result.extracted_data
            row = data_start_row + 1 + i
            ws.cell(row=row, column=1, value=data.get("date_of_issue", ""))
            ws.cell(row=row, column=2, value=self._parse_currency(data.get("amount_due", 0)))
            ws.cell(row=row, column=3, value=self._parse_currency(data.get("subtotal", 0)))
            ws.cell(row=row, column=4, value=self._parse_currency(data.get("tax", 0)))
            ws.cell(row=row, column=5, value=inv_result.invoice_number or "Unknown")

        data_end_row = data_start_row + len(report.invoice_results)

        # Column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

        if len(report.invoice_results) > 0:
            # Chart 1: Bar Chart - Invoice Amounts
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.style = 12
            bar_chart.title = "Invoice Amounts Over Time"
            bar_chart.y_axis.title = "Amount ($)"
            bar_chart.x_axis.title = "Date"

            data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
            cats_ref = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            bar_chart.width = 18
            bar_chart.height = 12

            ws.add_chart(bar_chart, "G3")

            # Chart 2: Line Chart - Trend
            line_chart = LineChart()
            line_chart.style = 13
            line_chart.title = "Invoice Amount Trend"
            line_chart.y_axis.title = "Amount ($)"
            line_chart.x_axis.title = "Invoice"

            data_ref2 = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
            cats_ref2 = Reference(ws, min_col=5, min_row=data_start_row + 1, max_row=data_end_row)
            line_chart.add_data(data_ref2, titles_from_data=True)
            line_chart.set_categories(cats_ref2)
            line_chart.width = 18
            line_chart.height = 12

            ws.add_chart(line_chart, "G20")

            # Chart 3: Stacked Bar - Subtotal vs Tax
            stacked_chart = BarChart()
            stacked_chart.type = "col"
            stacked_chart.grouping = "stacked"
            stacked_chart.style = 11
            stacked_chart.title = "Cost Breakdown (Subtotal + Tax)"
            stacked_chart.y_axis.title = "Amount ($)"

            data_ref3 = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row, max_col=4)
            cats_ref3 = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
            stacked_chart.add_data(data_ref3, titles_from_data=True)
            stacked_chart.set_categories(cats_ref3)
            stacked_chart.width = 18
            stacked_chart.height = 12

            ws.add_chart(stacked_chart, "A20")

            # Chart 4: Pie Chart - Distribution (if multiple invoices)
            if len(report.invoice_results) > 1:
                pie_chart = PieChart()
                pie_chart.title = "Invoice Amount Distribution"

                data_ref4 = Reference(ws, min_col=2, min_row=data_start_row + 1, max_row=data_end_row)
                cats_ref4 = Reference(ws, min_col=5, min_row=data_start_row + 1, max_row=data_end_row)
                pie_chart.add_data(data_ref4)
                pie_chart.set_categories(cats_ref4)
                pie_chart.width = 14
                pie_chart.height = 12

                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.dataLabels.showCatName = False

                ws.add_chart(pie_chart, "A37")

    def _populate_invoice_sheet(self, ws, inv_result: ValidationResult):
        """Populate a single invoice sheet."""
        data = inv_result.extracted_data
        current_row = 1

        # Invoice Header
        ws[f'A{current_row}'] = f"Invoice #{inv_result.invoice_number or 'Unknown'}"
        ws[f'A{current_row}'].font = Font(bold=True, size=16)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2

        # Basic Info Section
        ws[f'A{current_row}'] = "INVOICE DETAILS"
        ws[f'A{current_row}'].font = self.title_font
        current_row += 1

        basic_fields = [
            ("Invoice Number", "invoice_number"),
            ("Date of Issue", "date_of_issue"),
            ("Due Date", "due_date"),
            ("Billed To", "billed_to"),
            ("Email", "email"),
            ("Address", "address"),
            ("Workers", "workers"),
            ("Pictures Link", "pictures_link"),
        ]

        for label, key in basic_fields:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = data.get(key, "")
            current_row += 1

        current_row += 1

        # Line Items Section
        ws[f'A{current_row}'] = "LINE ITEMS"
        ws[f'A{current_row}'].font = self.title_font
        current_row += 1

        line_headers = ["Description", "Rate", "Qty", "Line Total"]
        for col, header in enumerate(line_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.thin_border
        current_row += 1

        line_items = data.get("line_items", [])
        for item in line_items:
            ws.cell(row=current_row, column=1, value=item.get("description", "")).border = self.thin_border
            rate_cell = ws.cell(row=current_row, column=2, value=item.get("rate", 0))
            rate_cell.number_format = self.currency_format
            rate_cell.border = self.thin_border
            ws.cell(row=current_row, column=3, value=item.get("qty", 0)).border = self.thin_border
            total_cell = ws.cell(row=current_row, column=4, value=item.get("line_total", 0))
            total_cell.number_format = self.currency_format
            total_cell.border = self.thin_border
            current_row += 1

        current_row += 1

        # Discounts Section
        discounts = data.get("discounts", [])
        if discounts:
            ws[f'A{current_row}'] = "DISCOUNTS"
            ws[f'A{current_row}'].font = self.title_font
            current_row += 1

            discount_headers = ["Discount Name", "Amount"]
            for col, header in enumerate(discount_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font_white
                cell.fill = self.header_fill
                cell.border = self.thin_border
            current_row += 1

            for discount in discounts:
                ws.cell(row=current_row, column=1, value=discount.get("name", "")).border = self.thin_border
                amount_cell = ws.cell(row=current_row, column=2, value=discount.get("amount", 0))
                amount_cell.number_format = self.currency_format
                amount_cell.border = self.thin_border
                current_row += 1

            current_row += 1

        # Totals Section
        ws[f'A{current_row}'] = "TOTALS"
        ws[f'A{current_row}'].font = self.title_font
        current_row += 1

        totals = [
            ("Subtotal", "subtotal"),
            ("Tax", "tax"),
            ("Total", "total"),
            ("Amount Paid", "amount_paid"),
            ("Amount Due", "amount_due"),
        ]

        for label, key in totals:
            if data.get(key) is not None:
                ws[f'A{current_row}'] = label
                ws[f'A{current_row}'].font = Font(bold=True)
                cell = ws[f'B{current_row}']
                cell.value = data.get(key, 0)
                cell.number_format = self.currency_format
                if key == "amount_due":
                    cell.font = Font(bold=True, size=12)
                current_row += 1

        current_row += 1

        # Validation Status
        ws[f'A{current_row}'] = "VALIDATION STATUS"
        ws[f'A{current_row}'].font = self.title_font
        current_row += 1

        ws[f'A{current_row}'] = "Status:"
        ws[f'A{current_row}'].font = Font(bold=True)
        status_cell = ws[f'B{current_row}']
        status_cell.value = "VALID" if inv_result.is_valid else "INVALID"
        status_cell.font = Font(bold=True, color="008000" if inv_result.is_valid else "FF0000")
        current_row += 1

        if inv_result.errors:
            ws[f'A{current_row}'] = "Errors:"
            ws[f'A{current_row}'].font = Font(bold=True, color="FF0000")
            current_row += 1
            for error in inv_result.errors:
                ws[f'A{current_row}'] = f"  - {error}"
                ws[f'A{current_row}'].font = Font(color="FF0000")
                current_row += 1

        if inv_result.warnings:
            ws[f'A{current_row}'] = "Warnings:"
            ws[f'A{current_row}'].font = Font(bold=True, color="FF8C00")
            current_row += 1
            for warning in inv_result.warnings:
                ws[f'A{current_row}'] = f"  - {warning}"
                ws[f'A{current_row}'].font = Font(color="FF8C00")
                current_row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15


# ============================================================================
# AI Invoice Validator
# ============================================================================

class AIInvoiceValidator:
    """Validates PDF invoices using Claude AI."""

    def __init__(self, api_key: str = None, templates_dir: str = None):
        self.api_key = api_key or os.environ.get("ANTHROPIC_API_KEY")
        if not self.api_key:
            raise ValueError("ANTHROPIC_API_KEY environment variable or api_key parameter required")
        self.client = anthropic.Anthropic(api_key=self.api_key)
        self.template_manager = TemplateManager(templates_dir)

    def _pdf_to_images(self, pdf_path: str, dpi: int = 150) -> list:
        """Convert PDF pages to base64-encoded images."""
        images = []
        doc = fitz.open(pdf_path)

        for page_num in range(len(doc)):
            page = doc[page_num]
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes("png")
            img_base64 = base64.standard_b64encode(img_bytes).decode("utf-8")
            images.append((page_num + 1, img_base64))

        doc.close()
        return images

    def _detect_vendor(self, images: list) -> Tuple[str, bool]:
        """
        Detect vendor from PDF images.
        Returns (vendor_name, is_known_vendor)
        """
        known_vendors = self.template_manager.get_all_vendor_names()

        # Build request to identify vendor
        content = [
            {
                "type": "text",
                "text": f"""Look at the first page of this document and identify the vendor/company name.

Known vendors in our system:
{json.dumps(known_vendors, indent=2)}

If the vendor matches one of the known vendors (even with slight spelling differences), return that exact name.
If it's a new vendor not in the list, return the vendor name as shown on the document.

Return ONLY the vendor name, nothing else."""
            },
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": images[0][1]  # First page only
                }
            }
        ]

        response = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=200,
            messages=[{"role": "user", "content": content}]
        )

        detected_vendor = response.content[0].text.strip()

        # Check if it matches a known vendor
        template = self.template_manager.get_template_by_vendor(detected_vendor)
        is_known = template is not None

        if is_known:
            # Return the canonical name from template
            return template["vendor"]["name"], True

        return detected_vendor, False

    def _create_template_from_pdf(self, images: list, vendor_name: str) -> dict:
        """Use AI to create a template from the PDF."""
        print(f"Creating new template for vendor: {vendor_name}")

        content = [
            {
                "type": "text",
                "text": f"""Analyze this invoice document and create a JSON template that describes its structure.

The vendor is: {vendor_name}

IMPORTANT: Use these exact normalized field names (snake_case) for required_fields:
- invoice_number (for Invoice #, Invoice Number)
- date_of_issue (for Invoice Date, Date of Issue)
- due_date (for Due Date)
- amount_due (for Amount Due)
- billed_to (for Bill To, Billed To)
- subtotal (for Sub Total, Subtotal)
- tax (for Tax, Sales Tax)
- total (for Total)

Create a comprehensive template with the following structure:
{{
  "template_name": "<Vendor Name> Invoice Template",
  "template_version": "1.0",
  "vendor": {{
    "name": "<exact vendor name>",
    "phone": "<phone number if visible>",
    "address": "<full address if visible>",
    "email": "<email if visible>",
    "website": "<website if visible>"
  }},
  "invoice_structure": {{
    "header_fields": ["invoice_number", "date_of_issue", "due_date", "amount_due"],
    "billing_fields": ["billed_to", "ship_to", "customer_id"],
    "line_item_fields": ["description", "unit_price", "quantity", "extended_price"],
    "discount_types": [],
    "total_fields": ["subtotal", "tax", "total", "amount_due"],
    "tax_info": {{
      "tax_name": "<name of tax, e.g., SRT, Sales Tax>",
      "tax_rate": <decimal rate, e.g., 0.036 for 3.6%>
    }}
  }},
  "validation_rules": {{
    "required_fields": ["invoice_number", "date_of_issue", "due_date", "amount_due", "billed_to", "subtotal", "tax", "total"],
    "date_format": "M/D/YYYY",
    "currency": "USD"
  }},
  "extraction_hints": {{
    "invoice_number_pattern": "<regex pattern>",
    "date_pattern": "\\\\d{{1,2}}/\\\\d{{1,2}}/\\\\d{{4}}",
    "amount_pattern": "\\\\$\\\\d+\\\\.\\\\d{{2}}",
    "multi_invoice_support": true,
    "pages_per_invoice": "1"
  }}
}}

Analyze the document carefully and fill in all the details you can observe.
Return ONLY the JSON template, no other text."""
            }
        ]

        # Add first few pages for analysis
        for page_num, img_base64 in images[:4]:
            content.append({
                "type": "text",
                "text": f"\n--- Page {page_num} ---"
            })
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": img_base64
                }
            })

        response = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            messages=[{"role": "user", "content": content}]
        )

        response_text = response.content[0].text

        # Extract JSON from response
        try:
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}') + 1
            if start_idx != -1 and end_idx > start_idx:
                json_str = response_text[start_idx:end_idx]
                template = json.loads(json_str)
                return template
        except json.JSONDecodeError as e:
            print(f"Warning: Could not parse template JSON: {e}")

        # Return a basic template if parsing fails
        return {
            "template_name": f"{vendor_name} Invoice Template",
            "template_version": "1.0",
            "vendor": {"name": vendor_name},
            "invoice_structure": {},
            "validation_rules": {"required_fields": ["invoice_number", "date_of_issue", "amount_due"]},
            "extraction_hints": {"multi_invoice_support": True}
        }

    def _extract_invoice_data(self, images: list, template: dict) -> list:
        """Extract invoice data from images using AI."""
        vendor_name = template.get("vendor", {}).get("name", "Unknown")
        tax_info = template.get("invoice_structure", {}).get("tax_info", {})
        tax_name = tax_info.get("tax_name", "Tax")

        content = [
            {
                "type": "text",
                "text": f"""Analyze these invoice images and extract data from each invoice found.

This PDF contains invoices from "{vendor_name}". Each invoice may span 1-2 pages.

For EACH invoice found, extract:
- invoice_number: The invoice number
- date_of_issue: Date of issue (format: MM/DD/YYYY)
- due_date: Due date (format: MM/DD/YYYY)
- billed_to: Customer/property name
- email: Billing email address
- address: Billing address
- amount_due: Final amount due (number only)
- subtotal: Subtotal before tax (number only)
- tax: {tax_name} amount (number only)
- total: Total amount (number only)
- amount_paid: Amount already paid (number only)
- line_items: Array of ALL line items with:
  - description: Full item description
  - rate: Unit rate (number)
  - qty: Quantity (number)
  - line_total: Line total (number)
- discounts: Array of ALL discounts with:
  - name: Discount name
  - amount: Discount amount (number)
- pictures_link: Any document/photo links
- workers: Worker assignments if shown
- page_numbers: Which page numbers contain this invoice

IMPORTANT: Extract ALL line items and ALL discounts completely.

Return ONLY a JSON array of invoice objects, no other text.

Example:
[
  {{
    "invoice_number": "12345",
    "date_of_issue": "01/15/2025",
    "due_date": "02/15/2025",
    "billed_to": "Customer Name",
    "email": "customer@email.com",
    "address": "123 Main St, City, State 12345",
    "amount_due": 500.00,
    "subtotal": 480.00,
    "tax": 20.00,
    "total": 500.00,
    "amount_paid": 0.00,
    "line_items": [
      {{"description": "Service description", "rate": 100.00, "qty": 1, "line_total": 100.00}}
    ],
    "discounts": [
      {{"name": "Discount Name", "amount": 10.00}}
    ],
    "pictures_link": "https://...",
    "workers": "@worker1 @worker2",
    "page_numbers": [1, 2]
  }}
]"""
            }
        ]

        # Add all images
        for page_num, img_base64 in images:
            content.append({
                "type": "text",
                "text": f"\n--- Page {page_num} ---"
            })
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": img_base64
                }
            })

        print("Sending images to Claude for extraction...")
        response = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=16000,
            messages=[{"role": "user", "content": content}]
        )

        response_text = response.content[0].text

        try:
            start_idx = response_text.find('[')
            end_idx = response_text.rfind(']') + 1
            if start_idx != -1 and end_idx > start_idx:
                json_str = response_text[start_idx:end_idx]
                invoices = json.loads(json_str)
                return invoices
        except json.JSONDecodeError as e:
            print(f"Warning: Could not parse JSON response: {e}")

        return []

    def _validate_invoice_data(self, invoice_data: dict, template: dict) -> ValidationResult:
        """Validate extracted invoice data against template rules."""
        page_numbers = invoice_data.get("page_numbers", [])
        result = ValidationResult(page_numbers=page_numbers)
        result.extracted_data = invoice_data.copy()

        # Mapping from template field names to extracted data field names
        field_mapping = {
            "invoice #": "invoice_number",
            "invoice number": "invoice_number",
            "invoice date": "date_of_issue",
            "date of issue": "date_of_issue",
            "due date": "due_date",
            "amount due": "amount_due",
            "bill to": "billed_to",
            "billed to": "billed_to",
            "sub total": "subtotal",
            "subtotal": "subtotal",
            "tax": "tax",
            "total": "total",
            "description": "line_items",
            "unit price": "line_items",
            "quantity": "line_items",
            "extended price": "line_items",
        }

        # Extract invoice number
        if invoice_data.get("invoice_number"):
            result.invoice_number = str(invoice_data["invoice_number"])
        else:
            result.add_error("Missing Invoice Number")

        # Validate required fields
        required_fields = template.get("validation_rules", {}).get(
            "required_fields",
            ["date_of_issue", "due_date", "amount_due", "billed_to"]
        )

        # Track which extracted fields we've already validated
        validated_extracted_fields = set()

        for field in required_fields:
            # Map template field name to extracted field name
            field_lower = field.lower()
            extracted_field = field_mapping.get(field_lower, field_lower.replace(" ", "_"))

            # Skip if we already validated this extracted field (e.g., multiple line_items fields)
            if extracted_field in validated_extracted_fields:
                continue
            validated_extracted_fields.add(extracted_field)

            # Check if field exists in extracted data
            value = invoice_data.get(extracted_field)
            if extracted_field == "line_items":
                # For line items, check if the list is non-empty
                if not value or len(value) == 0:
                    result.add_error(f"Missing required field: {field}")
            elif not value:
                result.add_error(f"Missing required field: {field}")

        # No date format validation or calculations - just extract and report as-is

        return result

    def validate_pdf(self, pdf_path: str) -> PDFValidationReport:
        """Main method to validate a PDF file."""
        pdf_path = Path(pdf_path)
        report = PDFValidationReport(filename=pdf_path.name)

        if not pdf_path.exists():
            report.global_errors.append(f"File not found: {pdf_path}")
            report.is_valid = False
            return report

        if pdf_path.suffix.lower() != '.pdf':
            report.global_errors.append(f"Not a PDF file: {pdf_path}")
            report.is_valid = False
            return report

        try:
            # Convert PDF to images
            print(f"Converting PDF to images...")
            images = self._pdf_to_images(str(pdf_path))
            report.total_pages = len(images)
            print(f"Converted {len(images)} pages to images")

            # Step 1: Detect vendor
            print("Detecting vendor...")
            vendor_name, is_known = self._detect_vendor(images)
            print(f"Detected vendor: {vendor_name} (Known: {is_known})")

            # Step 2: Get or create template
            if is_known:
                template = self.template_manager.get_template_by_vendor(vendor_name)
                report.template_name = template["vendor"]["name"]
                report.template_created = False
                print(f"Using existing template: {report.template_name}")
            else:
                # Create new template
                template = self._create_template_from_pdf(images, vendor_name)
                self.template_manager.save_template(template)
                report.template_name = vendor_name
                report.template_created = True
                print(f"Created new template for: {vendor_name}")

            # Step 3: Extract invoice data
            invoices = self._extract_invoice_data(images, template)

            if not invoices:
                report.global_errors.append("No invoices could be extracted from PDF")
                report.is_valid = False
                return report

            print(f"Found {len(invoices)} invoice(s) in PDF")

            # Step 4: Validate each invoice
            for invoice_data in invoices:
                result = self._validate_invoice_data(invoice_data, template)
                report.add_invoice_result(result)

        except anthropic.APIError as e:
            report.global_errors.append(f"Claude API error: {str(e)}")
            report.is_valid = False
        except Exception as e:
            import traceback
            report.global_errors.append(f"Error processing PDF: {str(e)}")
            traceback.print_exc()
            report.is_valid = False

        return report


# ============================================================================
# Report Formatting
# ============================================================================

def format_report(report: PDFValidationReport) -> str:
    """Format the validation report for display."""
    lines = []
    lines.append("=" * 70)
    lines.append(f"PDF VALIDATION REPORT: {report.filename}")
    lines.append("=" * 70)
    lines.append(f"Template: {report.template_name}")
    lines.append(f"Template Status: {'NEW (created)' if report.template_created else 'Existing'}")
    lines.append(f"Total Pages: {report.total_pages}")
    lines.append(f"Invoices Found: {report.invoices_found}")
    lines.append(f"Valid Invoices: {report.invoices_valid}")
    lines.append(f"Invalid Invoices: {report.invoices_invalid}")
    lines.append(f"Overall Status: {'VALID' if report.is_valid else 'INVALID'}")
    lines.append("")

    if report.global_errors:
        lines.append("GLOBAL ERRORS:")
        for error in report.global_errors:
            lines.append(f"  - {error}")
        lines.append("")

    for i, inv_result in enumerate(report.invoice_results, 1):
        lines.append("-" * 50)
        inv_num = inv_result.invoice_number or "Unknown"
        pages = ", ".join(map(str, inv_result.page_numbers)) if inv_result.page_numbers else "Unknown"
        status = "VALID" if inv_result.is_valid else "INVALID"
        lines.append(f"{i}. Invoice #{inv_num} (Pages: {pages}) - {status}")

        data = inv_result.extracted_data
        if data:
            lines.append("  Extracted Data:")
            key_fields = [
                "invoice_number", "date_of_issue", "due_date", "billed_to",
                "amount_due", "subtotal", "tax", "total"
            ]
            for key in key_fields:
                if key in data and data[key]:
                    value = data[key]
                    if isinstance(value, float):
                        lines.append(f"    {key}: ${value:.2f}")
                    else:
                        lines.append(f"    {key}: {value}")

            # Show ALL line items
            line_items = data.get("line_items", [])
            if line_items:
                lines.append(f"    line_items: {len(line_items)} item(s)")
                for item in line_items:
                    desc = item.get("description", "")
                    rate = item.get("rate", 0)
                    qty = item.get("qty", 0)
                    total = item.get("line_total", 0)
                    lines.append(f"      - {desc}")
                    lines.append(f"        Rate: ${rate:.2f} x Qty: {qty} = ${total:.2f}")

            # Show ALL discounts
            discounts = data.get("discounts", [])
            if discounts:
                lines.append(f"    discounts: {len(discounts)} discount(s)")
                for discount in discounts:
                    name = discount.get("name", "")
                    amount = discount.get("amount", 0)
                    lines.append(f"      - {name}: ${amount:.2f}")

        if inv_result.errors:
            lines.append("  Errors:")
            for error in inv_result.errors:
                lines.append(f"    - {error}")

        if inv_result.warnings:
            lines.append("  Warnings:")
            for warning in inv_result.warnings:
                lines.append(f"    - {warning}")
        lines.append("")

    lines.append("=" * 70)
    return "\n".join(lines)


# ============================================================================
# Public API
# ============================================================================

def validate_pdf(pdf_path: str, api_key: str = None, templates_dir: str = None) -> PDFValidationReport:
    """
    Validate a PDF file with automatic template detection/creation.

    Args:
        pdf_path: Path to the PDF file
        api_key: Anthropic API key (defaults to ANTHROPIC_API_KEY env var)
        templates_dir: Directory for templates (defaults to ./templates)

    Returns:
        PDFValidationReport with validation results
    """
    validator = AIInvoiceValidator(api_key, templates_dir)
    return validator.validate_pdf(pdf_path)


def export_to_excel(report: PDFValidationReport, output_path: str = None) -> str:
    """Export validation report to Excel."""
    exporter = ExcelExporter()
    return exporter.export_report_to_excel(report, output_path)


# ============================================================================
# Main Entry Point
# ============================================================================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python pdf_validator.py <pdf_path>")
        print("\nRequired:")
        print("  - Set ANTHROPIC_API_KEY environment variable")
        print("  - pip install anthropic PyMuPDF Pillow openpyxl")
        print("\nFeatures:")
        print("  - Automatically detects vendor from PDF")
        print("  - Creates new templates for unknown vendors")
        print("  - Validates invoices against template rules")
        print("  - Exports to Excel (one sheet per invoice)")
        print("\nTemplates stored in: ./templates/")
        sys.exit(1)

    pdf_path = sys.argv[1]

    # Check for API key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("Error: ANTHROPIC_API_KEY environment variable not set")
        print("Set it with: export ANTHROPIC_API_KEY='your-api-key'")
        sys.exit(1)

    # Validate PDF
    report = validate_pdf(pdf_path)

    # Print report to console
    print(format_report(report))

    # Export to Excel
    if report.invoices_found > 0:
        excel_path = export_to_excel(report)
        print(f"\nExcel file created: {excel_path}")

    sys.exit(0 if report.is_valid else 1)
